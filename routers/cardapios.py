"""
Menu.AI — Router de Cardápios
Gestão de cardápios gerados + workflow de aprovação.
"""
import html
import io
import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from typing import List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database.connection import get_db
from database.models import (
    Cardapio, CardapioDia, CardapioRefeicao,
    AprovacaoCardapio, JobAgente, Contrato, FichaTecnica
)
from database.schemas import (
    CardapioCreate, CardapioDetalhado, CardapioOut, CardapioUpdate,
    AprovacaoCreate, AprovacaoOut,
)
from routers.auth_supabase import exigir_role, get_usuario_atual
from services.knowledge_base import sync_cardapio_document
from services.knowledge_hooks import sync_knowledge_safe

router = APIRouter(prefix="/api/cardapios", tags=["Cardápios"])

TIPOS_REFEICAO_LABELS = {
    "cafe_manha": "Café da Manhã",
    "lanche_manha": "Lanche da Manhã",
    "almoco": "Almoço",
    "lanche_tarde": "Lanche da Tarde",
    "jantar": "Jantar",
    "ceia": "Ceia",
}


def _resolve_empresa_context(usuario, empresa_id: Optional[str]) -> str:
    requested = str(empresa_id).strip() if empresa_id else None
    user_empresa = str(usuario.empresa_id).strip() if getattr(usuario, "empresa_id", None) else None

    if usuario.role == "super_admin":
        resolved = requested or user_empresa
        if not resolved:
            raise HTTPException(
                status_code=400,
                detail="Super admin sem empresa no contexto. Informe empresa_id.",
            )
        return resolved

    if requested and requested != user_empresa:
        raise HTTPException(status_code=403, detail="empresa_id não corresponde ao utilizador autenticado.")
    if not user_empresa:
        raise HTTPException(status_code=400, detail="Utilizador sem empresa associada.")
    return user_empresa


def _norm_lookup(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _hydrate_cardapio_costs(db: Session, cardapio: Cardapio) -> bool:
    if not cardapio.empresa_id:
        return False
    refs = [ref for dia in (cardapio.dias or []) for ref in (dia.refeicoes or [])]
    if not refs:
        return False

    fichas = (
        db.query(FichaTecnica)
        .filter(FichaTecnica.empresa_id == cardapio.empresa_id, FichaTecnica.ativo == True)  # noqa: E712
        .all()
    )
    ficha_by_id = {str(f.id): f for f in fichas}
    ficha_by_code = {_norm_lookup(f.codigo): f for f in fichas if f.codigo}
    ficha_by_name = {_norm_lookup(f.nome): f for f in fichas if f.nome}

    changed = False
    total_geral = 0.0
    for dia in cardapio.dias or []:
        total_dia = 0.0
        for ref in dia.refeicoes or []:
            ficha = None
            if ref.ficha_tecnica_id:
                ficha = ficha_by_id.get(str(ref.ficha_tecnica_id))
            if ficha is None and ref.codigo_prato:
                ficha = ficha_by_code.get(_norm_lookup(ref.codigo_prato))
            if ficha is None and ref.nome_prato:
                ficha = ficha_by_name.get(_norm_lookup(ref.nome_prato))
            if ficha:
                if float(ficha.custo_porcao or 0.0) <= 0 and getattr(ficha, "ingredientes_ficha", None):
                    try:
                        from services.cascata import recalcular_ficha_unica

                        recalcular_ficha_unica(db, str(ficha.id))
                        db.flush()
                    except Exception:
                        pass
                if not ref.ficha_tecnica_id:
                    ref.ficha_tecnica_id = str(ficha.id)
                    changed = True
                if not ref.codigo_prato and ficha.codigo:
                    ref.codigo_prato = ficha.codigo
                    changed = True
                custo_ficha = round(float(ficha.custo_porcao or 0.0), 2)
                if custo_ficha > 0 and round(float(ref.custo_porcao or 0.0), 2) != custo_ficha:
                    ref.custo_porcao = custo_ficha
                    changed = True
            total_dia += float(ref.custo_porcao or 0.0)
        total_dia = round(total_dia, 2)
        if round(float(dia.custo_total or 0.0), 2) != total_dia:
            dia.custo_total = total_dia
            changed = True
        total_geral += total_dia
    media = round(total_geral / max(len(cardapio.dias or []), 1), 2)
    if round(float(cardapio.custo_medio_dia or 0.0), 2) != media:
        cardapio.custo_medio_dia = media
        changed = True
    if changed:
        cardapio.updated_at = datetime.utcnow()
    return changed


def _group_refeicoes_por_tipo(dia: CardapioDia) -> list[dict]:
    grupos: dict[str, list[CardapioRefeicao]] = defaultdict(list)
    for ref in sorted(dia.refeicoes or [], key=lambda item: ((item.ordem or 0), item.nome_prato or "")):
        grupos[str(ref.tipo_refeicao or "almoco")].append(ref)
    saida: list[dict] = []
    for tipo, items in grupos.items():
        componentes = []
        total = 0.0
        for ref in items:
            custo = round(float(ref.custo_porcao or 0.0), 2)
            total += custo
            componentes.append(
                {
                    "id": str(ref.id),
                    "tipo_refeicao": tipo,
                    "categoria": ref.categoria,
                    "ficha_tecnica_id": str(ref.ficha_tecnica_id) if ref.ficha_tecnica_id else None,
                    "codigo_prato": ref.codigo_prato,
                    "nome_prato": ref.nome_prato,
                    "custo_unitario": custo,
                    "custo_total_item": custo,
                    "observacoes": ref.observacoes,
                    "ordem": int(ref.ordem or 0),
                }
            )
        saida.append(
            {
                "tipo_refeicao": tipo,
                "label": TIPOS_REFEICAO_LABELS.get(tipo, tipo.replace("_", " ").title()),
                "custo_total": round(total, 2),
                "componentes": componentes,
            }
        )
    return saida


def _serialize_cardapio_detalhado(cardapio: Cardapio) -> dict:
    return {
        "id": str(cardapio.id),
        "empresa_id": str(cardapio.empresa_id),
        "contrato_id": str(cardapio.contrato_id) if cardapio.contrato_id else None,
        "criado_por_id": str(cardapio.criado_por_id) if cardapio.criado_por_id else None,
        "nome": cardapio.nome,
        "periodo_inicio": cardapio.periodo_inicio,
        "periodo_fim": cardapio.periodo_fim,
        "observacoes": cardapio.observacoes,
        "status": cardapio.status,
        "custo_medio_dia": cardapio.custo_medio_dia,
        "num_dias": cardapio.num_dias,
        "job_id": cardapio.job_id,
        "created_at": cardapio.created_at,
        "updated_at": cardapio.updated_at,
        "resultado_raw": cardapio.resultado_raw,
        "parametros_json": cardapio.parametros_json,
        "dias": [
            {
                "id": str(dia.id),
                "cardapio_id": str(dia.cardapio_id),
                "numero_dia": dia.numero_dia,
                "data": dia.data,
                "dia_semana": dia.dia_semana,
                "observacoes": dia.observacoes,
                "custo_total": round(float(dia.custo_total or 0.0), 2),
                "refeicoes": [
                    {
                        "id": str(ref.id),
                        "dia_id": str(ref.dia_id),
                        "tipo_refeicao": ref.tipo_refeicao,
                        "categoria": ref.categoria,
                        "codigo_prato": ref.codigo_prato,
                        "nome_prato": ref.nome_prato,
                        "custo_porcao": round(float(ref.custo_porcao or 0.0), 2),
                        "ficha_tecnica_id": str(ref.ficha_tecnica_id) if ref.ficha_tecnica_id else None,
                        "observacoes": ref.observacoes,
                        "ordem": int(ref.ordem or 0),
                    }
                    for ref in sorted(dia.refeicoes or [], key=lambda item: ((item.ordem or 0), item.nome_prato or ""))
                ],
                "refeicoes_agrupadas": _group_refeicoes_por_tipo(dia),
            }
            for dia in sorted(cardapio.dias or [], key=lambda item: (item.numero_dia or 0, item.data or datetime.utcnow().date()))
        ],
    }


@router.get("/", summary="Listar cardápios")
def listar(
    empresa_id: Optional[str] = Query(None),
    status_filtro: Optional[str] = Query(None, alias="status"),
    contrato_id: Optional[str] = Query(None),
    job_id: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual),
):
    q = db.query(Cardapio)

    eid = _resolve_empresa_context(usuario, empresa_id)
    q = q.filter(Cardapio.empresa_id == eid)

    if status_filtro:
        q = q.filter(Cardapio.status == status_filtro)
    if contrato_id:
        q = q.filter(Cardapio.contrato_id == contrato_id)
    if job_id:
        q = q.filter(Cardapio.job_id == job_id)

    total = q.count()
    items = q.order_by(Cardapio.created_at.desc()).offset(skip).limit(limit).all()
    return {
        "items": [CardapioOut.model_validate(c) for c in items],
        "total": total,
        "page": (skip // limit) + 1 if limit else 1,
        "per_page": limit,
        "pages": math.ceil(total / limit) if limit else 0,
    }


@router.get("/{cardapio_id}", response_model=CardapioDetalhado, summary="Buscar cardápio completo")
def buscar(
    cardapio_id: str,
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual),
):
    """Retorna cardápio completo com todos os dias e refeições."""
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")

    if usuario.role != "super_admin" and cardapio.empresa_id != usuario.empresa_id:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    if _hydrate_cardapio_costs(db, cardapio):
        db.commit()
        db.refresh(cardapio)

    return CardapioDetalhado.model_validate(_serialize_cardapio_detalhado(cardapio))


@router.post("/", response_model=CardapioOut, status_code=status.HTTP_201_CREATED,
             summary="Criar cardápio manualmente")
def criar(
    body: CardapioCreate,
    db: Session = Depends(get_db),
    usuario=Depends(exigir_role("super_admin", "admin", "nutricionista")),
):
    """Cria cardápio vazio para preenchimento manual."""
    empresa_id = body.empresa_id if usuario.role == "super_admin" else usuario.empresa_id
    cardapio = Cardapio(
        **body.model_dump(exclude={"empresa_id"}),
        empresa_id=empresa_id,
        criado_por_id=usuario.id,
    )
    db.add(cardapio)
    db.commit()
    db.refresh(cardapio)
    sync_knowledge_safe(sync_cardapio_document, db, cardapio)
    db.commit()
    return CardapioOut.model_validate(cardapio)


@router.patch("/{cardapio_id}", response_model=CardapioOut, summary="Atualizar cardápio")
def atualizar(
    cardapio_id: str,
    body: CardapioUpdate,
    db: Session = Depends(get_db),
    usuario=Depends(exigir_role("super_admin", "admin", "nutricionista")),
):
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")

    if usuario.role != "super_admin" and cardapio.empresa_id != usuario.empresa_id:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    for campo, valor in body.model_dump(exclude_unset=True).items():
        setattr(cardapio, campo, valor)

    cardapio.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(cardapio)
    sync_knowledge_safe(sync_cardapio_document, db, cardapio)
    db.commit()
    return CardapioOut.model_validate(cardapio)


@router.delete("/{cardapio_id}", status_code=status.HTTP_204_NO_CONTENT)
def arquivar(
    cardapio_id: str,
    db: Session = Depends(get_db),
    usuario=Depends(exigir_role("super_admin", "admin")),
):
    """Arquiva cardápio (soft delete)."""
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")
    cardapio.status = "arquivado"
    db.commit()
    db.refresh(cardapio)
    sync_knowledge_safe(sync_cardapio_document, db, cardapio)
    db.commit()


# ============================================================
# Aprovações (Workflow)
# ============================================================

@router.post("/{cardapio_id}/aprovacao", response_model=AprovacaoOut,
             status_code=status.HTTP_201_CREATED, summary="Aprovar ou reprovar cardápio")
def aprovar(
    cardapio_id: str,
    body: AprovacaoCreate,
    db: Session = Depends(get_db),
    usuario=Depends(exigir_role("super_admin", "admin", "gestor")),
):
    """
    Registra decisão de aprovação no workflow.
    - aprovado → cardápio vai para status 'aprovado'
    - reprovado → volta para 'em_revisao'
    - solicitado_revisao → status 'em_revisao'
    """
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")
    if body.cardapio_id and body.cardapio_id != cardapio_id:
        raise HTTPException(status_code=400, detail="cardapio_id do corpo não corresponde à URL.")

    aprovacao = AprovacaoCardapio(
        cardapio_id=cardapio_id,
        aprovado_por_id=usuario.id,
        status=body.status,
        comentario=body.comentario,
    )
    db.add(aprovacao)

    # Atualiza status do cardápio
    if body.status == "aprovado":
        cardapio.status = "aprovado"
    elif body.status in ("reprovado", "solicitado_revisao"):
        cardapio.status = "em_revisao"

    cardapio.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(aprovacao)
    db.refresh(cardapio)
    sync_knowledge_safe(sync_cardapio_document, db, cardapio)
    db.commit()
    return AprovacaoOut.model_validate(aprovacao)


@router.get("/{cardapio_id}/aprovacoes", response_model=List[AprovacaoOut],
            summary="Histórico de aprovações")
def historico_aprovacoes(
    cardapio_id: str,
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual),
):
    aprovacoes = db.query(AprovacaoCardapio).filter(
        AprovacaoCardapio.cardapio_id == cardapio_id
    ).order_by(AprovacaoCardapio.created_at.desc()).all()
    return [AprovacaoOut.model_validate(a) for a in aprovacoes]


@router.post("/{cardapio_id}/publicar", response_model=CardapioOut, summary="Publicar cardápio")
def publicar(
    cardapio_id: str,
    db: Session = Depends(get_db),
    usuario=Depends(exigir_role("super_admin", "admin", "gestor")),
):
    """Publica cardápio aprovado."""
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")
    if cardapio.status != "aprovado":
        raise HTTPException(status_code=400, detail="Cardápio precisa estar aprovado antes de publicar.")

    cardapio.status = "publicado"
    cardapio.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(cardapio)
    sync_knowledge_safe(sync_cardapio_document, db, cardapio)
    db.commit()
    return CardapioOut.model_validate(cardapio)


# ============================================================
# Exportação
# ============================================================

@router.get("/{cardapio_id}/exportar", summary="Exportar cardápio (XLSX, CSV, TXT ou PDF)")
def exportar(
    cardapio_id: str,
    formato: str = Query("xlsx", pattern=r"^(xlsx|csv|txt|pdf)$"),
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual),
):
    """Exporta cardápio em XLSX, CSV, TXT ou PDF."""
    cardapio = db.query(Cardapio).filter(Cardapio.id == cardapio_id).first()
    if not cardapio:
        raise HTTPException(status_code=404, detail="Cardápio não encontrado.")

    if not cardapio.resultado_raw and not cardapio.dias:
        raise HTTPException(status_code=404, detail="Cardápio sem conteúdo para exportar.")
    if _hydrate_cardapio_costs(db, cardapio):
        db.commit()
        db.refresh(cardapio)

    nome_arquivo = re.sub(r"[^\w\-]", "_", cardapio.nome)

    if formato == "txt":
        content = cardapio.resultado_raw or _gerar_txt(cardapio)
        return Response(
            content=content.encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.txt"'},
        )

    if formato == "csv":
        return _gerar_csv(cardapio, db, nome_arquivo)

    if formato == "pdf":
        buf = _gerar_pdf(cardapio, db)
        buf.seek(0)
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.pdf"'},
        )

    # XLSX multi-sheet com formatação
    buf = _gerar_xlsx(cardapio, db)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.xlsx"'},
    )


# ============================================================
# Exportação XLSX multi-sheet
# ============================================================

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_CAT_COLORS = {
    "ARROZ": "E2EFDA",
    "FEIJAO": "FCE4D6",
    "PRATO PROTEICO": "DCE6F1",
    "GUARNICAO": "FFF2CC",
    "SALADAS CRUA": "E4DFEC",
    "SALADA COZIDA": "E4DFEC",
    "SALADA ELABORADA": "E4DFEC",
    "ACOMPANHAMENTO": "D9D9D9",
    "SOBREMESA": "F8CBAD",
    "FRUTAS": "C6EFCE",
}


def _apply_header_style(ws, num_cols: int):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _apply_data_style(ws, num_rows: int, num_cols: int):
    for row in range(2, num_rows + 1):
        cat_cell = ws.cell(row=row, column=4)  # Coluna Categoria
        cat = (cat_cell.value or "").upper()
        fill_color = _CAT_COLORS.get(cat, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = _DATA_ALIGN
            cell.border = _THIN_BORDER
            cell.fill = row_fill


def _auto_width(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _write_sheet(ws, headers: list[str], rows: list[list], title: str = ""):
    ws.title = title[:31]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    _apply_header_style(ws, len(headers))
    _apply_data_style(ws, len(rows) + 1, len(headers))
    _auto_width(ws)


def _tem_refeicoes_estruturadas(cardapio: Cardapio) -> bool:
    for dia in cardapio.dias or []:
        if dia.refeicoes:
            return True
    return False


def _cardapio_para_dataframe(cardapio: Cardapio) -> pd.DataFrame:
    """Converte estrutura do banco para DataFrame (sem enriquecimento de custos)."""
    if cardapio.resultado_raw and "|" in cardapio.resultado_raw:
        return _extrair_markdown(cardapio.resultado_raw)

    if _tem_refeicoes_estruturadas(cardapio):
        rows = []
        for dia in cardapio.dias:
            meal_totals: dict[str, float] = defaultdict(float)
            for ref in dia.refeicoes:
                meal_totals[str(ref.tipo_refeicao or "almoco")] += float(ref.custo_porcao or 0.0)
            for ref in dia.refeicoes:
                # Custo: preserva valor real; marca None se zero para preenchimento posterior
                custo = ref.custo_porcao or 0
                rows.append({
                    "Dia": dia.numero_dia,
                    "Data": dia.data.strftime("%d/%m/%Y") if dia.data else "",
                    "Refeição": TIPOS_REFEICAO_LABELS.get(ref.tipo_refeicao, ref.tipo_refeicao.replace("_", " ").title()),
                    "Categoria": ref.categoria or "",
                    "Código": ref.codigo_prato or "",
                    "Prato": ref.nome_prato,
                    "Custo Item (R$)": round(custo, 2) if custo > 0 else None,
                    "Total Refeição (R$)": round(meal_totals.get(str(ref.tipo_refeicao or "almoco"), 0.0), 2),
                    "Total Dia (R$)": round(float(dia.custo_total or 0.0), 2),
                })
        return pd.DataFrame(rows) if rows else pd.DataFrame({"Cardápio": [cardapio.nome]})

    return pd.DataFrame({"Cardápio": [cardapio.nome]})


_CODIGO_COL_CANDIDATES = (
    "Código",
    "Codigo",
    "codigo",
    "Código Prato",
    "Codigo Prato",
    "Código da Ficha",
    "Cod",
)


def _coluna_codigo_df(df: pd.DataFrame) -> Optional[str]:
    for c in _CODIGO_COL_CANDIDATES:
        if c in df.columns:
            return c
    for col in df.columns:
        cl = str(col).strip().lower()
        if cl in ("código", "codigo", "cod", "código prato", "codigo prato"):
            return col
    return None


def _enriquecer_custos_dataframe(
    df: pd.DataFrame,
    cardapio: Cardapio,
    db: Session,
) -> pd.DataFrame:
    """Preenche ou sobrescreve coluna de custo do item com custo_porcao das fichas (match por código)."""
    if df.empty or len(df.columns) <= 1:
        return df
    col_cod = _coluna_codigo_df(df)
    if not col_cod or not cardapio.empresa_id:
        return df

    codigos_raw = df[col_cod].dropna().astype(str).str.strip()
    codigos_raw = codigos_raw[codigos_raw != ""]
    codigos = list({c for c in codigos_raw.unique() if c})
    if not codigos:
        return df

    fichas = (
        db.query(FichaTecnica)
        .filter(
            FichaTecnica.empresa_id == cardapio.empresa_id,
            FichaTecnica.codigo.in_(codigos),
        )
        .all()
    )
    custo_por_codigo = {f.codigo.strip(): float(f.custo_porcao or 0) for f in fichas}

    custo_col = "Custo Item (R$)" if "Custo Item (R$)" in df.columns else "Custo (R$)"
    if custo_col not in df.columns:
        df = df.copy()
        df[custo_col] = None

    def _custo_row(cod: object) -> Optional[float]:
        if cod is None or (isinstance(cod, float) and pd.isna(cod)):
            return None
        key = str(cod).strip()
        if not key:
            return None
        return custo_por_codigo.get(key)

    custos = df[col_cod].map(_custo_row)
    # Só preenche onde: (a) temos match na ficha E (b) o valor atual é nulo/zero
    existing = df[custo_col]
    needs_fill = existing.isna() | (existing == 0) | (existing == 0.0)
    mask = custos.notna() & needs_fill
    df = df.copy()
    df.loc[mask, custo_col] = custos[mask].round(2)
    return df


def _cardapio_export_dataframe(cardapio: Cardapio, db: Session) -> pd.DataFrame:
    """DataFrame para CSV/XLSX: sempre tenta enriquecer custos a partir das fichas."""
    df = _cardapio_para_dataframe(cardapio)
    # Sempre enriquece: preenche custos None/zero com dados reais das fichas técnicas
    return _limpar_dataframe_export(_enriquecer_custos_dataframe(df, cardapio, db))


def _norm_export_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _safe_day(value: object) -> Optional[int]:
    try:
        return int(float(str(value).replace("*", "").strip()))
    except (TypeError, ValueError):
        return None


_WEEKDAYS = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado", "domingo"]

_REFEICAO_TEMPLATE_ROWS = [
    ("PRATO PROTEICO", "Prato Proteico Principal", "% Consumo Principal"),
    ("PRATO PROTEICO", "Opção Proteica 2", "% Consumo Opção 2"),
    ("PRATO PROTEICO", "Opção Proteica 3", "% Consumo Opção 3"),
    ("ARROZ", "Arroz", "100"),
    ("FEIJAO", "Feijão", "100"),
    ("GUARNICAO", "Guarnição 1", "100"),
    ("GUARNICAO", "Guarnição 2", "100"),
    ("SALADA GRAOS", "Salada Grãos", "100"),
    ("SALADAS CRUA", "Salada Crua", "100"),
    ("SALADA COZIDA", "Salada Cozida", "100"),
    ("SALADA ELABORADA", "Salada Folhosa/Elaborada", "100"),
    ("SOBREMESA", "Sobremesa", "100"),
    ("BEBIDAS", "Bebida", "100"),
    ("FRUTAS", "Fruta", "100"),
    ("CUSTO GERENCIAL", "Custo Gerencial (R$)", ""),
]

_DESJEJUM_TEMPLATE_ROWS = [
    ("PAO", "Pão", "100"),
    ("RECHEIO", "Recheio 1", "% Consumo Recheio 1"),
    ("RECHEIO", "Recheio 2", "% Consumo Recheio 2"),
    ("ACOMPANHAMENTO", "Acompanhamento Café", "100"),
    ("BEBIDAS", "Bebida Café", "100"),
    ("FRUTAS", "Fruta Café", "100"),
    ("CUSTO GERENCIAL", "Custo Gerencial (R$)", ""),
]


def _df_rows_by_day_ref(df: pd.DataFrame) -> dict[tuple[int, str], dict[str, object]]:
    rows: dict[tuple[int, str], dict[str, object]] = {}
    if df.empty or "Dia" not in df.columns or "Refeição" not in df.columns:
        return rows
    for _, row in df.iterrows():
        day = _safe_day(row.get("Dia"))
        if day is None:
            continue
        rows[(day, _norm_export_key(row.get("Refeição")))] = row.to_dict()
    return rows


def _standard_value(row: Optional[dict[str, object]], value_col: str, pct_col: str) -> tuple[str, str]:
    if not row:
        return "", ""
    value = _limpar_texto_export(row.get(value_col, ""))
    if not value or value == "-":
        return "", ""
    if pct_col in row:
        pct = _limpar_texto_export(row.get(pct_col, ""))
    else:
        pct = pct_col
    if not pct or pct == "-":
        pct = ""
    return value, pct


def _style_standard_matrix(ws, start_row: int, end_row: int, max_col: int) -> None:
    title_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sub_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row_idx == start_row:
                cell.fill = title_fill
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
            elif row_idx in {start_row + 1, start_row + 2, start_row + 3, start_row + 4}:
                cell.fill = header_fill if row_idx < start_row + 4 else sub_fill
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF" if row_idx < start_row + 4 else "111827")
            elif col_idx == 1:
                cell.fill = sub_fill
                cell.font = Font(name="Calibri", bold=True, color="111827")


def _write_standard_matrix_sheet(
    wb,
    df: pd.DataFrame,
    *,
    title: str,
    refeicoes: list[str],
    template_rows: list[tuple[str, str, str]],
) -> bool:
    rows_by_key = _df_rows_by_day_ref(df)
    if not rows_by_key:
        return False

    ref_keys: dict[str, str] = {}
    for refeicao in refeicoes:
        ref_keys.setdefault(_norm_export_key(refeicao), refeicao)

    available_days = sorted({day for day, ref in rows_by_key if ref in ref_keys})
    if not available_days:
        return False

    ws = wb.create_sheet(title[:31])
    row_cursor = 1
    max_col = 15
    for ref_key, refeicao in ref_keys.items():
        days = sorted(day for day, key in rows_by_key if key == ref_key)
        if not days:
            continue
        for block_index in range(0, len(days), 7):
            week_days = days[block_index:block_index + 7]
            start = row_cursor
            ws.cell(row=row_cursor, column=1, value=f"{title} - {refeicao}")
            for col in range(2, max_col + 1):
                ws.cell(row=row_cursor, column=col, value="")
            row_cursor += 1

            header_rows = [
                ("Grupo de Pratos", "Sequência:"),
                ("Dia", None),
                ("Data", None),
                ("", "% Consumo"),
            ]
            for label, marker in header_rows:
                ws.cell(row=row_cursor, column=1, value=label)
                for i, day in enumerate(week_days):
                    value_col = 2 + i * 2
                    pct_col = value_col + 1
                    if label == "Grupo de Pratos":
                        ws.cell(row=row_cursor, column=value_col, value=marker)
                        ws.cell(row=row_cursor, column=pct_col, value=day)
                    elif label == "Dia":
                        ws.cell(row=row_cursor, column=value_col, value=_WEEKDAYS[(day - 1) % 7])
                        ws.cell(row=row_cursor, column=pct_col, value="NR")
                    elif label == "Data":
                        row = rows_by_key.get((day, ref_key), {})
                        ws.cell(row=row_cursor, column=value_col, value=_limpar_texto_export(row.get("Data", "")))
                        ws.cell(row=row_cursor, column=pct_col, value=_limpar_texto_export(row.get("Custo Gerencial (R$)", "")))
                    else:
                        ws.cell(row=row_cursor, column=value_col, value="")
                        ws.cell(row=row_cursor, column=pct_col, value=marker)
                row_cursor += 1

            for group, value_col_name, pct_col_name in template_rows:
                ws.cell(row=row_cursor, column=1, value=group)
                for i, day in enumerate(week_days):
                    row = rows_by_key.get((day, ref_key))
                    value, pct = _standard_value(row, value_col_name, pct_col_name)
                    value_col = 2 + i * 2
                    ws.cell(row=row_cursor, column=value_col, value=value)
                    ws.cell(row=row_cursor, column=value_col + 1, value=pct)
                row_cursor += 1

            _style_standard_matrix(ws, start, row_cursor - 1, max_col)
            row_cursor += 2

    if row_cursor == 1:
        wb.remove(ws)
        return False

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx % 2 == 0 else 11
    return True


def _extrair_markdown(texto: str) -> pd.DataFrame:
    linhas = [l.strip() for l in texto.split("\n") if "|" in l]
    linhas = [l for l in linhas if not re.match(r"^\|[\s\-:|]+\|$", l)]
    rows, header = [], None
    for l in linhas:
        stripped = l.strip().strip("|")
        cells = [c.strip() for c in stripped.split("|")]
        if not cells:
            continue
        if header is None:
            header = cells
        elif len(cells) >= len(header):
            rows.append(cells[: len(header)])
        elif len(cells) > 1:
            cells.extend([""] * (len(header) - len(cells)))
            rows.append(cells)
    if not header or not rows:
        return pd.DataFrame({"Cardápio": texto.split("\n")})
    return pd.DataFrame(rows, columns=header)


def _limpar_texto_export(valor: object) -> str:
    text = "" if valor is None else str(valor)
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"`([^`]*)`", r"\1", text)
    text = text.replace("\r", " ").replace("\n", " ")
    text = unicodedata.normalize("NFKC", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) not in {"So", "Cs", "Co"})
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _limpar_dataframe_export(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out.columns = [_limpar_texto_export(c) for c in out.columns]
    for col in out.columns:
        out[col] = out[col].map(_limpar_texto_export)
    return out


def _gerar_txt(cardapio: Cardapio) -> str:
    linhas = [f"CARDÁPIO: {cardapio.nome}\n"]
    for dia in cardapio.dias:
        linhas.append(f"\n=== Dia {dia.numero_dia} ===")
        for refeicao in _group_refeicoes_por_tipo(dia):
            linhas.append(f"  {refeicao['label']} — Total refeição: R$ {float(refeicao['custo_total'] or 0.0):.2f}")
            for comp in refeicao["componentes"]:
                linhas.append(
                    f"    [{comp.get('categoria') or '-'}] {comp.get('nome_prato') or '-'} — "
                    f"R$ {float(comp.get('custo_total_item') or 0.0):.2f}"
                )
        linhas.append(f"  Total do dia: R$ {dia.custo_total:.2f}")
    return "\n".join(linhas)


def _gerar_csv(cardapio: Cardapio, db: Session, nome_arquivo: str) -> Response:
    """Gera CSV com encoding UTF-8-BOM para compatibilidade com Excel."""
    df = _cardapio_export_dataframe(cardapio, db)
    buf = io.StringIO()
    df.to_csv(buf, index=False, sep=";", encoding="utf-8-sig")
    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.csv"'},
    )


def _gerar_pdf(cardapio: Cardapio, db: Session) -> io.BytesIO:
    """Gera PDF operacional usando a mesma matriz base do CSV/XLSX."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "MenuAITitle",
        parent=styles["Title"],
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#111827"),
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "MenuAISubtitle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=8,
    )
    header_style = ParagraphStyle(
        "MenuAITableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.8,
        leading=6.6,
        textColor=colors.white,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        "MenuAITableCell",
        parent=styles["Normal"],
        fontSize=5.6,
        leading=6.5,
        wordWrap="CJK",
    )

    def _p(value: object, style: ParagraphStyle = cell_style) -> Paragraph:
        return Paragraph(html.escape(_limpar_texto_export(value) or "-"), style)

    df = _cardapio_export_dataframe(cardapio, db)
    if df.empty:
        df = pd.DataFrame({"Cardápio": [cardapio.nome]})

    headers = [_limpar_texto_export(c) for c in df.columns]
    rows_data = [[_p(h, header_style) for h in headers]]
    for _, row in df.iterrows():
        rows_data.append([_p(row.get(col, "")) for col in df.columns])

    available_width = doc.width
    weights = []
    for h in headers:
        key = h.lower()
        if key in {"dia"}:
            weights.append(0.7)
        elif "refei" in key:
            weights.append(1.2)
        elif "tema" in key:
            weights.append(1.4)
        elif "custo" in key or "codigo" in key or "código" in key:
            weights.append(1.1)
        else:
            weights.append(2.1)
    total_weight = sum(weights) or 1
    col_widths = [available_width * (w / total_weight) for w in weights]

    story = [
        Paragraph(html.escape(_limpar_texto_export(cardapio.nome)), title_style),
        Paragraph(
            html.escape(
                f"Matriz operacional exportada em {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
            ),
            subtitle_style,
        ),
        Spacer(1, 4),
    ]

    table = Table(rows_data, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 5.8),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.2),
        ("TOPPADDING", (0, 0), (-1, -1), 2.4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.4),
    ]))
    story.append(table)

    def _footer(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#6b7280"))
        canvas.drawRightString(document.pagesize[0] - document.rightMargin, 0.35 * cm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf


def _gerar_xlsx(cardapio: Cardapio, db: Session) -> io.BytesIO:
    """Gera workbook XLSX com múltiplas sheets formatadas."""
    from openpyxl import Workbook

    wb = Workbook()

    # --- Sheet 1: Cardápio ---
    df = _cardapio_export_dataframe(cardapio, db)
    headers = list(df.columns)
    rows = [list(r) for r in df.values]
    ws1 = wb.active
    _write_sheet(ws1, headers, rows, "Cardápio")

    # --- Sheets no layout operacional enviado pelo cliente ---
    _write_standard_matrix_sheet(
        wb,
        df,
        title="Modelo Refeição",
        refeicoes=["Almoço", "Jantar"],
        template_rows=_REFEICAO_TEMPLATE_ROWS,
    )
    _write_standard_matrix_sheet(
        wb,
        df,
        title="Modelo Desjejum",
        refeicoes=["Café da manhã", "Cafe da Manha", "Desjejum", "Lanche da Manhã"],
        template_rows=_DESJEJUM_TEMPLATE_ROWS,
    )

    # --- Sheet 2: Lista de Compras (agrupar ingredientes das fichas usadas) ---
    compras_headers = ["Categoria", "Ingrediente", "Qtd Total (g)", "Custo Est. (R$)"]
    compras_rows = []
    if cardapio.dias:
        # Coletar pratos únicos do cardápio
        pratos = set()
        for dia in cardapio.dias:
            for ref in dia.refeicoes:
                if ref.codigo_prato:
                    pratos.add(ref.codigo_prato)
        # Buscar fichas e somar ingredientes
        from collections import defaultdict
        ing_map: dict[str, dict] = defaultdict(lambda: {"qtd": 0.0, "custo": 0.0})
        if pratos:
            fichas = db.query(FichaTecnica).filter(
                FichaTecnica.empresa_id == cardapio.empresa_id,
                FichaTecnica.codigo.in_(list(pratos)),
            ).all()
            for ficha in fichas:
                for item in ficha.ingredientes_ficha:
                    nome = item.ingrediente.nome if item.ingrediente else "Desconhecido"
                    ing_map[nome]["qtd"] += item.quantidade_bruta_g or 0
                    ing_map[nome]["custo"] += item.custo_calculado or 0
            for nome in sorted(ing_map):
                d = ing_map[nome]
                cat = ""
                for ficha in fichas:
                    for item in ficha.ingredientes_ficha:
                        if item.ingrediente and item.ingrediente.nome == nome:
                            cat = item.ingrediente.categoria or ""
                            break
                    if cat:
                        break
                compras_rows.append([cat, nome, round(d["qtd"], 1), round(d["custo"], 2)])
    _write_sheet(wb.create_sheet(), compras_headers, compras_rows, "Lista de Compras")

    # --- Sheet 3: Resumo por Dia ---
    resumo_dia_headers = ["Dia", "Data", "Nº Refeições", "Custo Total (R$)", "Custo Médio/Refeição"]
    resumo_dia_rows = []
    custo_total_geral = 0.0
    for dia in cardapio.dias:
        n_refs = len(_group_refeicoes_por_tipo(dia))
        custo_dia = dia.custo_total or sum(r.custo_porcao or 0 for r in dia.refeicoes)
        custo_medio = round(custo_dia / n_refs, 2) if n_refs else 0
        resumo_dia_rows.append([
            dia.numero_dia,
            dia.data.strftime("%d/%m/%Y") if dia.data else "",
            n_refs,
            round(custo_dia, 2),
            custo_medio,
        ])
        custo_total_geral += custo_dia
    resumo_dia_rows.append([
        "", "TOTAL", "", round(custo_total_geral, 2),
        round(custo_total_geral / max(len(resumo_dia_rows), 1), 2),
    ])
    _write_sheet(wb.create_sheet(), resumo_dia_headers, resumo_dia_rows, "Resumo Diário")

    # --- Sheet 4: Resumo Nutricional (por categoria) ---
    nut_headers = ["Categoria", "Nº Pratos", "Custo Médio (R$)"]
    nut_rows = []
    if cardapio.dias:
        cat_stats: dict[str, dict] = defaultdict(lambda: {"n": 0, "custo": 0.0})
        for dia in cardapio.dias:
            for ref in dia.refeicoes:
                cat = ref.categoria or "Sem categoria"
                cat_stats[cat]["n"] += 1
                cat_stats[cat]["custo"] += ref.custo_porcao or 0
        for cat in sorted(cat_stats):
            s = cat_stats[cat]
            nut_rows.append([cat, s["n"], round(s["custo"] / max(s["n"], 1), 2)])
    _write_sheet(wb.create_sheet(), nut_headers, nut_rows, "Resumo Nutricional")

    # --- Sheet 5: Incidência Proteica ---
    prot_headers = ["Tipo Proteína", "Ocorrências", "% do Total", "Custo Médio (R$)"]
    prot_rows = []
    if cardapio.dias:
        prot_stats: dict[str, dict] = defaultdict(lambda: {"n": 0, "custo": 0.0})
        total_prots = 0
        for dia in cardapio.dias:
            for ref in dia.refeicoes:
                cat_upper = (ref.categoria or "").upper()
                if "PROT" in cat_upper:
                    subtipo = _detectar_subtipo_proteina(ref.nome_prato)
                    prot_stats[subtipo]["n"] += 1
                    prot_stats[subtipo]["custo"] += ref.custo_porcao or 0
                    total_prots += 1
        for tipo in sorted(prot_stats):
            s = prot_stats[tipo]
            pct = round((s["n"] / max(total_prots, 1)) * 100, 1)
            prot_rows.append([
                tipo,
                s["n"],
                f"{pct}%",
                round(s["custo"] / max(s["n"], 1), 2),
            ])
    _write_sheet(wb.create_sheet(), prot_headers, prot_rows, "Incidência Proteica")

    buf = io.BytesIO()
    wb.save(buf)
    return buf


# ============================================================
# Helpers de classificação proteica
# ============================================================

_SUBTIPOS_PROTEINA = {
    "Bovino": ["bovino", "carne", "patinho", "acém", "músculo", "alcatra",
               "lagarto", "cupim", "maminha", "contra filé", "contrafilé",
               "coxão", "paleta", "costela bovina"],
    "Frango": ["frango", "ave", "peito de frango", "coxa", "sobrecoxa",
               "filé de frango", "frango assado", "frango grelhado"],
    "Suíno": ["suíno", "porco", "lombo", "pernil", "bisteca",
              "costela suína", "linguiça"],
    "Peixe": ["peixe", "tilápia", "merluza", "sardinha", "atum",
              "bacalhau", "pescada", "salmão", "filé de peixe"],
    "Ovo": ["ovo", "omelete", "ovos", "ovo cozido", "ovo mexido"],
    "Vegetal": ["soja", "grão-de-bico", "grão de bico", "lentilha",
                "feijão", "tofu", "PTS", "proteína de soja",
                "proteina texturizada"],
}


def _detectar_subtipo_proteina(nome_prato: str) -> str:
    """Classifica o prato proteico por subtipo baseado no nome."""
    nome_lower = (nome_prato or "").lower()
    for subtipo, keywords in _SUBTIPOS_PROTEINA.items():
        if any(kw.lower() in nome_lower for kw in keywords):
            return subtipo
    return "Outro"
